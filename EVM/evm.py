"""
EVM - Electronic Voting Machine
School Election System
Supports: Keyboard Mode (now) | Arduino Mode (future)
With Candidate Photo Support & Backward Compatibility
"""

import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog
import sqlite3
import json
import os
import sys
import hashlib
from datetime import datetime
import threading
import shutil

# ── Optional: Image support ────────────────────────────────────────────────
try:
    from PIL import Image, ImageTk
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False

# ── Optional: Excel export ────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "election.json")
DB_PATH     = os.path.join(BASE_DIR, "election.db")
EXPORT_PATH = os.path.join(BASE_DIR, "Election_Results.xlsx")
IMAGES_DIR  = os.path.join(BASE_DIR, "images")

# Logo path - try both PNG and JPG
LOGO_PATH = None
for logo_file in ["school_logo.jpg", "school_logo.png", "school_logo.jpeg"]:
    logo_candidate = os.path.join(IMAGES_DIR, logo_file)
    if os.path.exists(logo_candidate):
        LOGO_PATH = logo_candidate
        break

# Create directories if missing
os.makedirs(IMAGES_DIR, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "exports"), exist_ok=True)

# ── Colour palette ─────────────────────────────────────────────────────────
C = {
    "bg":        "#0D1B2A",   # deep navy
    "card":      "#1B2A3B",   # slightly lighter navy
    "accent":    "#F4A261",   # warm amber
    "accent2":   "#E76F51",   # coral
    "green":     "#2A9D8F",   # confirm green
    "green_hov": "#1E7D71",
    "white":     "#FFFFFF",
    "muted":     "#8DA9C4",
    "border":    "#2E4057",
    "red":       "#E63946",
    "admin_bg":  "#14213D",
    "selected":  "#264653",
    "sel_border":"#F4A261",
}

FONTS = {
    "title":    ("Segoe UI", 28, "bold"),
    "subtitle": ("Segoe UI", 14),
    "post":     ("Segoe UI", 22, "bold"),
    "cand":     ("Segoe UI", 16),
    "cand_key": ("Courier New", 18, "bold"),
    "btn":      ("Segoe UI", 14, "bold"),
    "small":    ("Segoe UI", 11),
    "admin":    ("Segoe UI", 13),
    "mono":     ("Courier New", 13),
}

# ════════════════════════════════════════════════════════════════════════════
#  HELPER: Extract candidate name (backward compatible)
# ════════════════════════════════════════════════════════════════════════════
def candidate_name(candidate):
    """
    Extract candidate name from either string or dict format.
    Supports both legacy and new formats:
    - String: "Candidate A"
    - Dict: {"name": "Candidate A", "image": "images/a.jpg"}
    """
    if isinstance(candidate, dict):
        return candidate.get("name", "")
    return candidate


def get_candidate_image(candidate):
    """
    Extract image path from candidate dict.
    Returns None if candidate is string or no image.
    """
    if isinstance(candidate, dict):
        return candidate.get("image")
    return None


# ════════════════════════════════════════════════════════════════════════════
#  DATABASE
# ════════════════════════════════════════════════════════════════════════════
class Database:
    def __init__(self, path):
        self.path = path
        self._init()

    def _conn(self):
        c = sqlite3.connect(self.path)
        c.execute("PRAGMA journal_mode=WAL")
        return c

    def _init(self):
        c = self._conn()
        c.execute("""
            CREATE TABLE IF NOT EXISTS votes (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                voter_seq INTEGER,
                post      TEXT NOT NULL,
                candidate TEXT NOT NULL,
                timestamp TEXT NOT NULL
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                start_ts  TEXT NOT NULL,
                end_ts    TEXT
            )
        """)
        c.commit()
        c.close()

    def start_session(self):
        c = self._conn()
        cur = c.execute("INSERT INTO sessions (start_ts) VALUES (?)",
                        (datetime.now().isoformat(),))
        c.commit()
        sid = cur.lastrowid
        c.close()
        return sid

    def close_session(self, sid):
        c = self._conn()
        c.execute("UPDATE sessions SET end_ts=? WHERE id=?",
                  (datetime.now().isoformat(), sid))
        c.commit()
        c.close()

    def next_voter_seq(self):
        c = self._conn()
        row = c.execute("SELECT MAX(voter_seq) FROM votes").fetchone()
        c.close()
        return (row[0] or 0) + 1

    def save_vote(self, voter_seq, post, candidate_name_str):
        """Save vote with candidate name (always string, never dict)"""
        c = self._conn()
        c.execute(
            "INSERT INTO votes (voter_seq,post,candidate,timestamp) VALUES(?,?,?,?)",
            (voter_seq, post, candidate_name_str, datetime.now().isoformat())
        )
        c.commit()
        c.close()

    def get_results(self):
        """Returns {post: {candidate: count}}"""
        c = self._conn()
        rows = c.execute(
            "SELECT post, candidate, COUNT(*) FROM votes GROUP BY post, candidate"
        ).fetchall()
        c.close()
        results = {}
        for post, cand, cnt in rows:
            results.setdefault(post, {})[cand] = cnt
        return results

    def get_audit_log(self):
        c = self._conn()
        rows = c.execute(
            "SELECT id, voter_seq, timestamp, post, candidate FROM votes ORDER BY id"
        ).fetchall()
        c.close()
        return rows

    def total_voters(self):
        c = self._conn()
        row = c.execute("SELECT COUNT(DISTINCT voter_seq) FROM votes").fetchone()
        c.close()
        return row[0] or 0

    def reset_election(self):
        c = self._conn()
        c.execute("DELETE FROM votes")
        c.execute("DELETE FROM sessions")
        c.commit()
        c.close()

    def backup(self):
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = os.path.join(BASE_DIR, f"backup_{ts}.db")
        shutil.copy2(self.path, dst)
        return dst


# ════════════════════════════════════════════════════════════════════════════
#  CONFIG LOADER
# ════════════════════════════════════════════════════════════════════════════
def load_config():
    if not os.path.exists(CONFIG_PATH):
        # create default config
        default = {
            "election_title": "SCHOOL ELECTION 2026",
            "admin_password": "admin123",
            "posts": {
                "Head Boy":  ["Rahul Sharma", "Arjun Das", "Rohan Gupta"],
                "Head Girl": ["Priya Das",    "Sneha Roy", "Ananya Singh"]
            }
        }
        with open(CONFIG_PATH, "w") as f:
            json.dump(default, f, indent=2)
    with open(CONFIG_PATH) as f:
        return json.load(f)


def save_config(config):
    """Save config to election.json"""
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=2)


# ════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORTER (with backward compatible candidate names)
# ════════════════════════════════════════════════════════════════════════════
def export_excel(db: Database, config: dict, path: str) -> bool:
    if not EXCEL_AVAILABLE:
        return False

    wb  = openpyxl.Workbook()
    results = db.get_results()

    # ── Results sheet ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Results"

    hdr_fill  = PatternFill("solid", fgColor="1B2A3B")
    hdr_font  = Font(bold=True, color="F4A261", size=13)
    sub_fill  = PatternFill("solid", fgColor="264653")
    sub_font  = Font(bold=True, color="FFFFFF", size=12)
    body_font = Font(size=11)
    thin = Side(style="thin", color="2E4057")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12

    row = 1
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws[f"A{row}"]
    cell.value    = config.get("election_title", "SCHOOL ELECTION 2026")
    cell.font     = Font(bold=True, color="F4A261", size=16)
    cell.fill     = PatternFill("solid", fgColor="0D1B2A")
    cell.alignment= ctr
    row += 1

    for post_name, candidates in config["posts"].items():
        # Post header
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws[f"A{row}"]
        cell.value     = post_name.upper()
        cell.font      = sub_font
        cell.fill      = sub_fill
        cell.alignment = ctr
        row += 1

        # Column headers
        for col, label in enumerate(["Post", "Candidate", "Votes"], 1):
            c = ws.cell(row=row, column=col, value=label)
            c.font      = hdr_font
            c.fill      = hdr_fill
            c.alignment = ctr
            c.border    = bdr
        row += 1

        post_votes = results.get(post_name, {})
        # Convert candidates to names and sort
        ranked = sorted(candidates,
                        key=lambda x: post_votes.get(candidate_name(x), 0), reverse=True)
        for cand in ranked:
            cand_name_str = candidate_name(cand)
            votes = post_votes.get(cand_name_str, 0)
            for col, val in enumerate([post_name, cand_name_str, votes], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = body_font
                c.alignment = ctr
                c.border    = bdr
            row += 1
        row += 1  # blank between posts

    # ── Audit sheet ───────────────────────────────────────────────
    wa = wb.create_sheet("Audit Log")
    wa.column_dimensions["A"].width = 8
    wa.column_dimensions["B"].width = 12
    wa.column_dimensions["C"].width = 22
    wa.column_dimensions["D"].width = 22
    wa.column_dimensions["E"].width = 22

    for col, label in enumerate(["ID", "Voter#", "Timestamp", "Post", "Candidate"], 1):
        c = wa.cell(row=1, column=col, value=label)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = ctr
        c.border    = bdr

    for r, row_data in enumerate(db.get_audit_log(), 2):
        for col, val in enumerate(row_data, 1):
            c = wa.cell(row=r, column=col, value=val)
            c.font   = body_font
            c.border = bdr

    # ── Stats sheet ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Statistics")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    stats = [
        ("Total Voters",    db.total_voters()),
        ("Total Votes Cast",len(db.get_audit_log())),
        ("Number of Posts", len(config["posts"])),
        ("Export Time",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for r, (k, v) in enumerate(stats, 1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True, size=12)
        ws2.cell(row=r, column=2, value=v).font = Font(size=12)

    wb.save(path)
    return True


# ════════════════════════════════════════════════════════════════════════════
#  IMAGE LOADER (with error handling)
# ════════════════════════════════════════════════════════════════════════════
def load_image(image_path, size=(80, 80)):
    """
    Load and resize image from path.
    Returns PhotoImage or None if image not found/corrupted.
    """
    if not PILLOW_AVAILABLE:
        return None
    
    try:
        if not os.path.exists(image_path):
            return None
        
        img = Image.open(image_path)
        img.thumbnail(size, Image.Resampling.LANCZOS)
        # Convert to PhotoImage
        return ImageTk.PhotoImage(img)
    except Exception as e:
        # Silently fail for corrupted images
        return None


# ════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ════════════════════════════════════════════════════════════════════════════
class EVMApp:
    KEYMAP = ["1", "2", "3", "4", "5"]

    def __init__(self, root: tk.Tk):
        self.root   = root
        self.config = load_config()
        self.db     = Database(DB_PATH)
        self.posts  = list(self.config["posts"].items())  # [(post, [cands])]
        self.input_mode = "keyboard"  # "keyboard" | "arduino"
        self.photo_cache = []  # Prevent Tkinter images from being garbage collected
        self.logo_photo = None  # Keep reference to logo image

        self._setup_window()
        self._build_ui()
        self._bind_keys()
        self.show_welcome()

    # ── Window setup ───────────────────────────────────────────────────────
    def _setup_window(self):
        title = self.config.get("election_title", "SCHOOL ELECTION 2026")
        self.root.title(title)
        self.root.configure(bg=C["bg"])
        self.root.attributes("-fullscreen", True)
        self.root.resizable(False, False)
        # Allow Escape only in admin (blocked in voting)
        self.root.protocol("WM_DELETE_WINDOW", self._safe_close)

    def _safe_close(self):
        # Do nothing during voting; handled via admin
        pass

    # ── Key binding (Keyboard Mode) ────────────────────────────────────────
    def _bind_keys(self):
        for key in self.KEYMAP:
            self.root.bind(key,
                lambda e, k=key: self._on_key(k))
        self.root.bind("<Return>",   lambda e: self._on_confirm())
        self.root.bind("0",          lambda e: self._on_start())

    def _on_key(self, key):
        """Called when 1/2/3/4/5 is pressed."""
        if self.screen == "voting":
            idx = self.KEYMAP.index(key)
            self._select_candidate(idx)

    def _on_confirm(self):
        if self.screen == "voting":
            self._confirm_vote()

    def _on_start(self):
        if self.screen in ("welcome", "thankyou"):
            self._start_voting()

    # ── Arduino hook (future) ──────────────────────────────────────────────
    def arduino_input(self, signal: str):
        """
        Call this method from an Arduino listener thread.
        signal: "1","2","3","4","5" for candidates, "CONFIRM" to confirm.
        Example (future):
            app.arduino_input("1")
            app.arduino_input("CONFIRM")
        """
        signal = signal.strip()
        if signal in self.KEYMAP:
            idx = self.KEYMAP.index(signal)
            self.root.after(0, lambda: self._select_candidate(idx))
        elif signal.upper() == "CONFIRM":
            self.root.after(0, self._confirm_vote)
        elif signal.upper() == "START":
            self.root.after(0, self._start_voting)

    # ═══════════════════════════════════════════════════════════════════════
    #  UI BUILDER
    # ═══════════════════════════════════════════════════════════════════════
    def _build_ui(self):
        # ── Top bar (always visible) ────────────────────────────────────────
        self.topbar = tk.Frame(self.root, bg=C["card"], height=54)
        self.topbar.pack(fill="x", side="top")
        self.topbar.pack_propagate(False)

        title_text = self.config.get("election_title", "SCHOOL ELECTION 2026")
        tk.Label(self.topbar, text=title_text,
                 bg=C["card"], fg=C["accent"],
                 font=("Segoe UI", 16, "bold")).pack(side="left", padx=20)

        self.clock_label = tk.Label(self.topbar, text="",
                                    bg=C["card"], fg=C["muted"],
                                    font=FONTS["small"])
        self.clock_label.pack(side="right", padx=20)
        self._tick_clock()

        # Admin button (small, top-right)
        tk.Button(self.topbar, text="⚙ ADMIN",
                  bg=C["admin_bg"], fg=C["muted"],
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", cursor="hand2",
                  command=self._open_admin,
                  padx=8, pady=2).pack(side="right", padx=8)

        # ── Main content area ──────────────────────────────────────────────
        self.main = tk.Frame(self.root, bg=C["bg"])
        self.main.pack(fill="both", expand=True)

        self.screen = None
        self.current_post_idx  = 0
        self.current_selection = None   # index into candidates list
        self.voter_seq = 1
        self.session_id = None
        self.session_votes = {}  # {post: candidate_name}

    def _tick_clock(self):
        self.clock_label.config(
            text=datetime.now().strftime("  %d %b %Y   %H:%M:%S  "))
        self.root.after(1000, self._tick_clock)

    def _clear_main(self):
        for w in self.main.winfo_children():
            w.destroy()
        # Clear photo cache when screen changes
        self.photo_cache = []

    def _load_logo(self, size=(120, 120)):
        """Load school logo image"""
        if not PILLOW_AVAILABLE or not LOGO_PATH:
            return None
        
        try:
            if not os.path.exists(LOGO_PATH):
                return None
            
            img = Image.open(LOGO_PATH)
            img.thumbnail(size, Image.Resampling.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    # ═══════════════════════════════════════════════════════════════════════
    #  SCREENS
    # ═══════════════════════════════════════════════════════════════════════

    # ── Welcome Screen ─────────────────────────────────────────────────────
    def show_welcome(self):
        self._clear_main()
        self.screen = "welcome"

        # Logo at top center
        logo_frame = tk.Frame(self.main, bg=C["bg"])
        logo_frame.pack(pady=(20, 10))
        
        logo_photo = self._load_logo(size=(120, 120))
        if logo_photo:
            self.logo_photo = logo_photo
            logo_label = tk.Label(logo_frame, image=logo_photo, bg=C["bg"])
            logo_label.pack()

        center = tk.Frame(self.main, bg=C["bg"])
        center.place(relx=0.5, rely=0.5, anchor="center")

        title = self.config.get("election_title", "SCHOOL ELECTION 2026")
        tk.Label(center,
                 text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                 bg=C["bg"], fg=C["border"],
                 font=("Courier New", 14)).pack()

        tk.Label(center, text=f"  {title}  ",
                 bg=C["bg"], fg=C["accent"],
                 font=("Segoe UI", 36, "bold")).pack(pady=(6, 6))

        tk.Label(center,
                 text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                 bg=C["bg"], fg=C["border"],
                 font=("Courier New", 14)).pack()

        tk.Label(center, text="Your vote is your voice.",
                 bg=C["bg"], fg=C["muted"],
                 font=("Segoe UI", 16, "italic")).pack(pady=(24, 40))

        start_btn = tk.Button(center,
                              text="▶   START VOTING",
                              bg=C["accent"], fg=C["bg"],
                              font=("Segoe UI", 20, "bold"),
                              relief="flat", cursor="hand2",
                              padx=50, pady=18,
                              command=self._start_voting)
        start_btn.pack()
        self._hover(start_btn, C["accent2"], C["accent"])

        tk.Label(center, text="(or press  0)",
                 bg=C["bg"], fg=C["muted"],
                 font=FONTS["small"]).pack(pady=(12, 0))

        # Stats strip at bottom
        voters = self.db.total_voters()
        tk.Label(self.main,
                 text=f"  Voters who have voted today: {voters}",
                 bg=C["card"], fg=C["muted"],
                 font=FONTS["small"]).place(relx=0, rely=1.0, anchor="sw",
                                            relwidth=1)

    # ── Voting Screen ──────────────────────────────────────────────────────
    def _start_voting(self):
        self.current_post_idx  = 0
        self.current_selection = None
        self.session_votes = {}
        self.voter_seq  = self.db.next_voter_seq()
        self.session_id = self.db.start_session()
        self._show_post()

    def _show_post(self):
        self._clear_main()
        self.screen = "voting"
        self.current_selection = None

        post_name, candidates = self.posts[self.current_post_idx]
        total_posts = len(self.posts)

        # Progress bar
        progress_frame = tk.Frame(self.main, bg=C["card"], height=6)
        progress_frame.pack(fill="x")
        pct = (self.current_post_idx) / total_posts
        tk.Frame(progress_frame,
                 bg=C["accent"], height=6,
                 width=int(self.root.winfo_screenwidth() * pct)
                 ).pack(side="left")

        # Top section with logo on left and nav on right
        top_section = tk.Frame(self.main, bg=C["bg"])
        top_section.pack(fill="x", padx=20, pady=(10, 0))

        # Logo on top left
        logo_photo = self._load_logo(size=(80, 80))
        if logo_photo:
            self.logo_photo = logo_photo
            logo_label = tk.Label(top_section, image=logo_photo, bg=C["bg"])
            logo_label.pack(side="left")

        # Post counter on right
        nav_bar = tk.Frame(top_section, bg=C["bg"])
        nav_bar.pack(side="right")
        tk.Label(nav_bar,
                 text=f"Post  {self.current_post_idx+1}  of  {total_posts}",
                 bg=C["bg"], fg=C["muted"],
                 font=FONTS["small"]).pack()

        # Post name
        post_label = tk.Frame(self.main, bg=C["card"], pady=12)
        post_label.pack(fill="x", padx=60, pady=(0, 18))
        tk.Label(post_label, text=post_name.upper(),
                 bg=C["card"], fg=C["white"],
                 font=FONTS["post"]).pack()

        tk.Label(self.main,
                 text="Press the number next to your choice, then press ENTER to confirm",
                 bg=C["bg"], fg=C["muted"],
                 font=FONTS["small"]).pack(pady=(0, 10))

        # Candidate buttons
        self.cand_buttons = []
        self.cand_frames  = []
        cand_area = tk.Frame(self.main, bg=C["bg"])
        cand_area.pack(pady=8)

        for i, cand in enumerate(candidates):
            key = self.KEYMAP[i]
            cand_name_str = candidate_name(cand)
            image_path = get_candidate_image(cand)
            
            frame = tk.Frame(cand_area,
                             bg=C["card"],
                             bd=2, relief="flat",
                             cursor="hand2")
            frame.pack(fill="x", padx=80, pady=5, ipady=4)

            inner = tk.Frame(frame, bg=C["card"])
            inner.pack(fill="x", padx=8, pady=4)

            key_lbl = tk.Label(inner,
                               text=f" [{key}] ",
                               bg=C["accent"], fg=C["bg"],
                               font=FONTS["cand_key"],
                               width=4)
            key_lbl.pack(side="left", padx=(0, 16))

            # Load and display candidate photo if available
            if image_path and os.path.exists(os.path.join(BASE_DIR, image_path)):
                try:
                    photo = load_image(os.path.join(BASE_DIR, image_path), size=(80, 80))
                    if photo:
                        img_lbl = tk.Label(inner, image=photo, bg=C["card"])
                        img_lbl.image = photo  # Keep reference
                        self.photo_cache.append(photo)
                        img_lbl.pack(side="left", padx=(0, 12))
                except Exception:
                    pass  # Silently skip if image can't be loaded

            name_lbl = tk.Label(inner,
                                text=cand_name_str,
                                bg=C["card"], fg=C["white"],
                                font=FONTS["cand"],
                                anchor="w")
            name_lbl.pack(side="left", fill="x", expand=True)

            # bind click
            for widget in (frame, inner, key_lbl, name_lbl):
                widget.bind("<Button-1>",
                            lambda e, idx=i: self._select_candidate(idx))

            self.cand_frames.append(frame)
            self.cand_buttons.append((frame, inner, name_lbl, key_lbl))

        # Selection feedback
        self.selection_label = tk.Label(self.main, text="",
                                        bg=C["bg"], fg=C["accent"],
                                        font=("Segoe UI", 15, "bold"))
        self.selection_label.pack(pady=(16, 6))

        # Confirm button
        self.confirm_btn = tk.Button(self.main,
                                     text="✔   CONFIRM VOTE   (Enter)",
                                     bg=C["border"], fg=C["muted"],
                                     font=("Segoe UI", 16, "bold"),
                                     relief="flat", cursor="hand2",
                                     padx=40, pady=14,
                                     state="disabled",
                                     command=self._confirm_vote)
        self.confirm_btn.pack(pady=(4, 0))

    def _select_candidate(self, idx):
        post_name, candidates = self.posts[self.current_post_idx]
        if idx >= len(candidates):
            return

        self.current_selection = idx

        # Reset all frames
        for frame, inner, name_lbl, key_lbl in self.cand_buttons:
            frame.config(bg=C["card"])
            inner.config(bg=C["card"])
            name_lbl.config(bg=C["card"])

        # Highlight selected
        frame, inner, name_lbl, key_lbl = self.cand_buttons[idx]
        frame.config(bg=C["selected"])
        inner.config(bg=C["selected"])
        name_lbl.config(bg=C["selected"], fg=C["accent"])

        selected_name = candidate_name(candidates[idx])
        self.selection_label.config(
            text=f"✔  You selected:  {selected_name}"
        )
        self.confirm_btn.config(
            state="normal",
            bg=C["green"], fg=C["white"],
            cursor="hand2"
        )
        self._hover(self.confirm_btn, C["green_hov"], C["green"])

    def _confirm_vote(self):
        if self.current_selection is None:
            return
        post_name, candidates = self.posts[self.current_post_idx]
        chosen = candidates[self.current_selection]
        
        # Extract name from candidate (handles both string and dict)
        chosen_name = candidate_name(chosen)

        self.session_votes[post_name] = chosen_name
        self.db.save_vote(self.voter_seq, post_name, chosen_name)

        # Next post or finish
        self.current_post_idx += 1
        if self.current_post_idx < len(self.posts):
            self._show_post()
        else:
            self._show_thankyou()

    # ── Beep helper ────────────────────────────────────────────────────────
    def _beep(self):
        """Play a confirmation beep in a background thread (non-blocking)."""
        def _play():
            try:
                import winsound                          # Windows
                winsound.Beep(1000, 200)                 # 1000 Hz, 200 ms
                import time; time.sleep(0.15)
                winsound.Beep(1000, 200)
                time.sleep(0.15)
                winsound.Beep(1320, 400)                 # higher note to finish
            except ImportError:
                try:
                    import subprocess, sys               # Linux / macOS
                    if sys.platform == "darwin":
                        subprocess.Popen(
                            ["afplay", "/System/Library/Sounds/Glass.aiff"])
                    else:
                        result = subprocess.run(
                            ["beep", "-f", "1000", "-l", "200",
                             "-D", "100", "-f", "1000", "-l", "200",
                             "-D", "100", "-f", "1320", "-l", "400"],
                            capture_output=True
                        )
                        if result.returncode != 0:
                            raise FileNotFoundError
                except Exception:
                    # Last resort: ASCII bell via print
                    print("", end="", flush=True)
        threading.Thread(target=_play, daemon=True).start()

    # ── Thank-you Screen ───────────────────────────────────────────────────
    def _show_thankyou(self):
        self._clear_main()
        self.screen = "thankyou"
        self.db.close_session(self.session_id)
        self._beep()

        # Logo at top center
        logo_frame = tk.Frame(self.main, bg=C["bg"])
        logo_frame.pack(pady=(20, 10))
        
        logo_photo = self._load_logo(size=(120, 120))
        if logo_photo:
            self.logo_photo = logo_photo
            logo_label = tk.Label(logo_frame, image=logo_photo, bg=C["bg"])
            logo_label.pack()

        center = tk.Frame(self.main, bg=C["bg"])
        center.place(relx=0.5, rely=0.5, anchor="center")

        tk.Label(center,
                 text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                 bg=C["bg"], fg=C["border"],
                 font=("Courier New", 14)).pack()

        tk.Label(center, text="✔  THANK YOU FOR VOTING  ✔",
                 bg=C["bg"], fg=C["green"],
                 font=("Segoe UI", 32, "bold")).pack(pady=(10, 6))

        tk.Label(center,
                 text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                 bg=C["bg"], fg=C["border"],
                 font=("Courier New", 14)).pack()

        tk.Label(center, text="Your votes have been recorded.",
                 bg=C["bg"], fg=C["muted"],
                 font=("Segoe UI", 15)).pack(pady=(16, 4))

        # Summary of this voter's choices
        summary = tk.Frame(center, bg=C["card"], padx=30, pady=14)
        summary.pack(pady=(18, 24))
        tk.Label(summary, text="Your Choices",
                 bg=C["card"], fg=C["accent"],
                 font=("Segoe UI", 13, "bold")).pack()
        for post, cand_name in self.session_votes.items():
            tk.Label(summary,
                     text=f"  {post}:  {cand_name}",
                     bg=C["card"], fg=C["white"],
                     font=("Segoe UI", 12)).pack(anchor="w")

        tk.Label(center,
                 text="Voter #" + str(self.voter_seq),
                 bg=C["bg"], fg=C["muted"],
                 font=FONTS["small"]).pack(pady=(0, 20))

        next_btn = tk.Button(center,
                             text="▶   NEXT VOTER  (press 0)",
                             bg=C["accent"], fg=C["bg"],
                             font=("Segoe UI", 18, "bold"),
                             relief="flat", cursor="hand2",
                             padx=40, pady=14,
                             command=self.show_welcome)
        next_btn.pack()
        self._hover(next_btn, C["accent2"], C["accent"])

    # ═══════════════════════════════════════════════════════════════════════
    #  ADMIN PANEL
    # ═══════════════════════════════════════════════════════════════════════
    def _open_admin(self):
        pwd = simpledialog.askstring(
            "Admin Login", "Enter Admin Password:",
            parent=self.root, show="*")
        if pwd is None:
            return
        correct = self.config.get("admin_password", "admin123")
        if pwd != correct:
            messagebox.showerror("Access Denied", "Wrong password.", parent=self.root)
            return
        self._show_admin_panel()

    def _show_admin_panel(self):
        win = tk.Toplevel(self.root)
        win.title("Admin Panel")
        win.configure(bg=C["admin_bg"])
        win.geometry("700x620")
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win,
                 text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                 bg=C["admin_bg"], fg=C["border"],
                 font=("Courier New", 12)).pack(pady=(14, 0))
        tk.Label(win, text="ADMIN PANEL",
                 bg=C["admin_bg"], fg=C["accent"],
                 font=("Segoe UI", 22, "bold")).pack()
        tk.Label(win,
                 text="━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                 bg=C["admin_bg"], fg=C["border"],
                 font=("Courier New", 12)).pack(pady=(0, 14))

        btn_opts = dict(font=("Segoe UI", 14, "bold"),
                        relief="flat", cursor="hand2",
                        padx=20, pady=10, width=28)

        def make_btn(text, color, cmd):
            b = tk.Button(win, text=text, bg=color,
                          fg=C["white" if color != C["accent"] else "bg"],
                          command=cmd, **btn_opts)
            if color == C["accent"]:
                b.config(fg=C["bg"])
            b.pack(pady=4)
            return b

        make_btn("📊  View Results",      C["green"],    lambda: self._admin_results(win))
        make_btn("📥  Export Excel",       C["accent"],   lambda: self._admin_export(win))
        make_btn("📈  Election Statistics",C["selected"], lambda: self._admin_stats(win))
        make_btn("🖼  Manage Candidate Photos", C["muted"], lambda: self._manage_photos(win))
        make_btn("🗑   Reset Election",     C["red"],      lambda: self._admin_reset(win))
        make_btn("💾  Backup Database",    C["border"],   lambda: self._admin_backup(win))

        tk.Button(win, text="✕   Exit Admin",
                  bg=C["admin_bg"], fg=C["muted"],
                  font=("Segoe UI", 12), relief="flat",
                  cursor="hand2", command=win.destroy).pack(pady=(16, 0))

    # ── View Results ───────────────────────────────────────────────────────
    def _admin_results(self, parent):
        win = tk.Toplevel(parent)
        win.title("Live Results")
        win.configure(bg=C["bg"])
        win.geometry("600x620")
        win.grab_set()

        tk.Label(win, text="LIVE RESULTS",
                 bg=C["bg"], fg=C["accent"],
                 font=("Segoe UI", 20, "bold")).pack(pady=(18, 4))

        results = self.db.get_results()
        canvas = tk.Canvas(win, bg=C["bg"], highlightthickness=0)
        scroll = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=20)

        inner = tk.Frame(canvas, bg=C["bg"])
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>",
                   lambda e: canvas.configure(
                       scrollregion=canvas.bbox("all")))

        for post_name, candidates in self.config["posts"].items():
            tk.Label(inner,
                     text=f"  {post_name.upper()}",
                     bg=C["card"], fg=C["accent"],
                     font=("Segoe UI", 14, "bold"),
                     anchor="w").pack(fill="x", pady=(12, 2))

            post_votes = results.get(post_name, {})
            # Convert candidates to names for sorting and voting lookup
            cand_names = [candidate_name(c) for c in candidates]
            max_votes = max((post_votes.get(name, 0) for name in cand_names), default=1)

            for cand in candidates:
                cand_name_str = candidate_name(cand)
                votes = post_votes.get(cand_name_str, 0)
                pct   = votes / max_votes if max_votes else 0

                row = tk.Frame(inner, bg=C["bg"])
                row.pack(fill="x", pady=2)

                tk.Label(row, text=f"{cand_name_str:<28}",
                         bg=C["bg"], fg=C["white"],
                         font=FONTS["admin"], width=28, anchor="w").pack(side="left")

                # bar
                bar_bg = tk.Frame(row, bg=C["border"], height=20, width=200)
                bar_bg.pack(side="left", padx=(4, 6))
                bar_bg.pack_propagate(False)
                bar_fill = tk.Frame(bar_bg,
                                    bg=C["accent"] if pct == 1.0 else C["green"],
                                    height=20,
                                    width=int(200 * pct))
                bar_fill.place(x=0, y=0)

                tk.Label(row, text=str(votes),
                         bg=C["bg"], fg=C["white"],
                         font=FONTS["mono"], width=5).pack(side="left")

        tk.Button(win, text="Close",
                  bg=C["border"], fg=C["white"],
                  font=FONTS["small"], relief="flat",
                  command=win.destroy, padx=20).pack(pady=12)

    # ── Export Excel ───────────────────────────────────────────────────────
    def _admin_export(self, parent):
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error",
                "openpyxl not installed.\nRun: pip install openpyxl",
                parent=parent)
            return
        try:
            export_excel(self.db, self.config, EXPORT_PATH)
            messagebox.showinfo("Export Complete",
                f"File saved:\n{EXPORT_PATH}", parent=parent)
        except Exception as ex:
            messagebox.showerror("Export Failed", str(ex), parent=parent)

    # ── Statistics ─────────────────────────────────────────────────────────
    def _admin_stats(self, parent):
        win = tk.Toplevel(parent)
        win.title("Election Statistics")
        win.configure(bg=C["bg"])
        win.geometry("420x380")
        win.grab_set()

        tk.Label(win, text="ELECTION STATISTICS",
                 bg=C["bg"], fg=C["accent"],
                 font=("Segoe UI", 18, "bold")).pack(pady=(20, 10))

        results = self.db.get_results()
        total_voters = self.db.total_voters()
        total_votes  = sum(
            sum(v.values()) for v in results.values()
        )

        stats = [
            ("Total Voters",      total_voters),
            ("Total Votes Cast",  total_votes),
            ("Number of Posts",   len(self.config["posts"])),
        ]
        for post_name, candidates in self.config["posts"].items():
            post_votes = results.get(post_name, {})
            # Find leader using candidate names
            winner = max(
                (candidate_name(c) for c in candidates),
                key=lambda x: post_votes.get(x, 0),
                default="No votes yet"
            )
            stats.append((f"  {post_name} Leader", winner))

        frame = tk.Frame(win, bg=C["card"], padx=24, pady=16)
        frame.pack(padx=30, pady=8, fill="x")
        for key, val in stats:
            row = tk.Frame(frame, bg=C["card"])
            row.pack(fill="x", pady=3)
            tk.Label(row, text=key,
                     bg=C["card"], fg=C["muted"],
                     font=("Segoe UI", 12), anchor="w",
                     width=26).pack(side="left")
            tk.Label(row, text=str(val),
                     bg=C["card"], fg=C["white"],
                     font=("Segoe UI", 12, "bold")).pack(side="left")

        tk.Button(win, text="Close",
                  bg=C["border"], fg=C["white"],
                  font=FONTS["small"], relief="flat",
                  command=win.destroy, padx=20).pack(pady=14)

    # ── Manage Candidate Photos ────────────────────────────────────────────
    def _manage_photos(self, parent):
        win = tk.Toplevel(parent)
        win.title("Manage Candidate Photos")
        win.configure(bg=C["bg"])
        win.geometry("500x400")
        win.grab_set()

        tk.Label(win, text="MANAGE CANDIDATE PHOTOS",
                 bg=C["bg"], fg=C["accent"],
                 font=("Segoe UI", 18, "bold")).pack(pady=(20, 10))

        # Post selection
        tk.Label(win, text="Select Post:",
                 bg=C["bg"], fg=C["white"],
                 font=("Segoe UI", 12)).pack(pady=(10, 4))
        
        posts_list = list(self.config["posts"].keys())
        post_var = tk.StringVar(value=posts_list[0] if posts_list else "")
        post_menu = tk.OptionMenu(win, post_var, *posts_list)
        post_menu.config(bg=C["card"], fg=C["white"], font=("Segoe UI", 11))
        post_menu.pack(pady=4)

        # Candidate selection (will be updated when post changes)
        tk.Label(win, text="Select Candidate:",
                 bg=C["bg"], fg=C["white"],
                 font=("Segoe UI", 12)).pack(pady=(10, 4))
        
        cand_var = tk.StringVar()
        cand_menu = tk.OptionMenu(win, cand_var, "")
        cand_menu.config(bg=C["card"], fg=C["white"], font=("Segoe UI", 11))
        cand_menu.pack(pady=4)

        def update_candidates(post_name):
            candidates = self.config["posts"].get(post_name, [])
            cand_names = [candidate_name(c) for c in candidates]
            menu = cand_menu["menu"]
            menu.delete(0, "end")
            for name in cand_names:
                menu.add_command(label=name, command=lambda v=name: cand_var.set(v))
            if cand_names:
                cand_var.set(cand_names[0])

        def on_post_change(*args):
            update_candidates(post_var.get())

        post_var.trace("write", on_post_change)
        update_candidates(post_var.get())

        # Image file selection button
        def select_image():
            post_name = post_var.get()
            cand_name = cand_var.get()
            
            if not post_name or not cand_name:
                messagebox.showerror("Error", "Please select both post and candidate.", parent=win)
                return

            file_path = filedialog.askopenfilename(
                parent=win,
                title="Select Candidate Image",
                filetypes=[("Image files", "*.jpg *.jpeg *.png *.gif"), ("All files", "*.*")]
            )
            
            if not file_path:
                return

            try:
                # Copy image to images directory
                filename = os.path.basename(file_path)
                dest_path = os.path.join(IMAGES_DIR, filename)
                shutil.copy2(file_path, dest_path)

                # Find candidate in config and update it
                candidates = self.config["posts"][post_name]
                for i, cand in enumerate(candidates):
                    if candidate_name(cand) == cand_name:
                        # Update candidate to dict format if it isn't already
                        if isinstance(cand, str):
                            self.config["posts"][post_name][i] = {
                                "name": cand,
                                "image": f"images/{filename}"
                            }
                        else:
                            self.config["posts"][post_name][i]["image"] = f"images/{filename}"
                        break

                # Save config and reload
                save_config(self.config)
                self.posts = list(self.config["posts"].items())
                
                messagebox.showinfo("Success", 
                    f"Image saved for {cand_name}:\n{dest_path}", 
                    parent=win)
            except Exception as ex:
                messagebox.showerror("Error", f"Failed to save image:\n{str(ex)}", parent=win)

        tk.Button(win, text="📁  Select Image",
                  bg=C["accent"], fg=C["bg"],
                  font=("Segoe UI", 12, "bold"),
                  relief="flat", cursor="hand2",
                  command=select_image, padx=20, pady=10).pack(pady=(20, 10))

        tk.Button(win, text="Close",
                  bg=C["border"], fg=C["white"],
                  font=FONTS["small"], relief="flat",
                  command=win.destroy, padx=20).pack(pady=10)

    # ── Reset Election ─────────────────────────────────────────────────────
    def _admin_reset(self, parent):
        if not messagebox.askyesno(
                "Confirm Reset",
                "This will DELETE all votes permanently.\nAre you sure?",
                parent=parent):
            return
        confirm2 = simpledialog.askstring(
            "Final Confirmation",
            "Type  RESET  to confirm:",
            parent=parent)
        if confirm2 == "RESET":
            self.db.reset_election()
            messagebox.showinfo("Done", "All votes have been reset.", parent=parent)
        else:
            messagebox.showinfo("Cancelled", "Reset cancelled.", parent=parent)

    # ── Backup ──────────────────────────────────────────────────────────────
    def _admin_backup(self, parent):
        try:
            path = self.db.backup()
            messagebox.showinfo("Backup Created",
                f"Backup saved:\n{path}", parent=parent)
        except Exception as ex:
            messagebox.showerror("Backup Failed", str(ex), parent=parent)

    # ═══════════════════════════════════════════════════════════════════════
    #  HELPERS
    # ═══════════════════════════════════════════════════════════════════════
    @staticmethod
    def _hover(btn, on_color, off_color):
        btn.bind("<Enter>", lambda e: btn.config(bg=on_color))
        btn.bind("<Leave>", lambda e: btn.config(bg=off_color))


# ════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    app  = EVMApp(root)

    # ── Arduino future hook ────────────────────────────────────────────────
    # To enable Arduino mode, start a thread here that reads from serial port
    # and calls app.arduino_input("A"), app.arduino_input("CONFIRM"), etc.
    #
    # Example (uncomment and install pyserial when ready):
    #
    # import serial, threading
    # def arduino_listener():
    #     with serial.Serial('/dev/ttyACM0', 9600) as ser:
    #         while True:
    #             line = ser.readline().decode().strip()
    #             app.arduino_input(line)
    # threading.Thread(target=arduino_listener, daemon=True).start()

    root.mainloop()


if __name__ == "__main__":
    main()
